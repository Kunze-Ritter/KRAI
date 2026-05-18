"""Unit tests for TicketIngestionProcessor."""

from unittest.mock import AsyncMock

import pytest

from backend.processors.ticket_ingestion_processor import TicketIngestionProcessor


@pytest.fixture
def mock_db_adapter():
    """Create mock database adapter."""
    adapter = AsyncMock()
    adapter.execute_query = AsyncMock(return_value=None)
    return adapter


@pytest.fixture
def processor(mock_db_adapter):
    """Create processor instance."""
    return TicketIngestionProcessor(mock_db_adapter)


@pytest.mark.asyncio
async def test_ingest_km_anfragen_of_success(processor, mock_km_excel_of, mock_db_adapter):
    """Test successful ingestion of KM Office tickets."""
    mock_db_adapter.execute_query.return_value = None
    processor.db_adapter = mock_db_adapter

    count = await processor.ingest_file(mock_km_excel_of, "km_anfragen_of")

    assert count == 100
    assert mock_db_adapter.execute_query.call_count == 100


@pytest.mark.asyncio
async def test_ingest_km_anfragen_pp(processor, mock_km_excel_pp, mock_db_adapter):
    """Test ingestion of KM Production tickets."""
    mock_db_adapter.execute_query.return_value = None
    processor.db_adapter = mock_db_adapter

    count = await processor.ingest_file(mock_km_excel_pp, "km_anfragen_pp")

    assert count == 50


@pytest.mark.asyncio
async def test_ingest_km_anfragen_sol(processor, mock_km_excel_sol, mock_db_adapter):
    """Test ingestion of KM Solutions tickets."""
    mock_db_adapter.execute_query.return_value = None
    processor.db_adapter = mock_db_adapter

    count = await processor.ingest_file(mock_km_excel_sol, "km_anfragen_sol")

    assert count == 30


@pytest.mark.asyncio
async def test_ingest_invalid_source_system(processor, mock_km_excel_of):
    """Test that invalid source system raises error."""
    with pytest.raises(ValueError) as exc_info:
        await processor.ingest_file(mock_km_excel_of, "invalid_source")

    assert "Invalid source_system" in str(exc_info.value)


@pytest.mark.asyncio
async def test_ingest_missing_excel_file(processor):
    """Test that missing file raises error."""
    with pytest.raises(ValueError) as exc_info:
        await processor.ingest_file(None, "km_anfragen_of")

    assert "Missing excel_file" in str(exc_info.value)


@pytest.mark.asyncio
async def test_extract_tickets_with_error_codes(processor, mock_km_excel_of):
    """Test that error codes are properly extracted."""
    import openpyxl

    excel_file = mock_km_excel_of
    workbook = openpyxl.load_workbook(excel_file, data_only=True)
    worksheet = workbook.active

    tickets = await processor._extract_tickets(worksheet, "km_anfragen_of")

    # Check first ticket
    assert len(tickets) > 0
    first_ticket = tickets[0]
    assert first_ticket["source_system"] == "km_anfragen_of"
    assert isinstance(first_ticket["error_codes"], list)
    assert len(first_ticket["error_codes"]) > 0


@pytest.mark.asyncio
async def test_extract_tickets_with_parts(processor, mock_km_excel_of):
    """Test that replaced parts are properly extracted."""
    import openpyxl

    excel_file = mock_km_excel_of
    workbook = openpyxl.load_workbook(excel_file, data_only=True)
    worksheet = workbook.active

    tickets = await processor._extract_tickets(worksheet, "km_anfragen_of")

    first_ticket = tickets[0]
    assert "replaced_parts" in first_ticket
    assert isinstance(first_ticket["replaced_parts"], list)


@pytest.mark.asyncio
async def test_parse_date_formats(processor):
    """Test various date format parsing."""
    dates = [
        "2024-05-15",
        "15.05.2024",
        "2024-05-15 14:30:00",
        "15.05.2024 14:30:00",
    ]

    for date_str in dates:
        parsed = processor._parse_date(date_str)
        assert parsed is not None, f"Failed to parse {date_str}"


@pytest.mark.asyncio
async def test_repair_time_parsing(processor):
    """Test repair time integer parsing."""
    from io import BytesIO

    import openpyxl

    # Create test workbook with various repair times
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(
        [
            "TicketID",
            "Erstellt",
            "Gerät/Modell",
            "Problem Kurz",
            "Problem Lang",
            "Lösung",
            "Fehlercodes",
            "Ersatzteile",
            "Reparaturzeit (min)",
        ]
    )
    ws.append(["T001", "2024-05-15", "KM C308", "Test", "Test", "Test", "", "", "15"])
    ws.append(["T002", "2024-05-15", "KM C308", "Test", "Test", "Test", "", "", "30.5"])  # float
    ws.append(["T003", "2024-05-15", "KM C308", "Test", "Test", "Test", "", "", "invalid"])  # invalid

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    workbook = openpyxl.load_workbook(output, data_only=True)
    worksheet = workbook.active
    tickets = await processor._extract_tickets(worksheet, "test")

    assert tickets[0]["repair_time_minutes"] == 15
    # Second row should be skipped due to float, or parsed as int
    assert tickets[1]["repair_time_minutes"] in [30, 30.5, None]
    assert tickets[2]["repair_time_minutes"] is None  # Invalid value


def test_get_supported_sources(processor):
    """Test that processor knows all supported sources."""
    assert "km_anfragen_of" in processor.SUPPORTED_SOURCES
    assert "km_anfragen_pp" in processor.SUPPORTED_SOURCES
    assert "km_anfragen_sol" in processor.SUPPORTED_SOURCES
