"""Fixtures for PM tests."""

import io
from datetime import datetime, timedelta

import openpyxl
import pytest


@pytest.fixture
def mock_km_excel_of():
    """Create mock KM Office (OF) Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tickets"

    # Header
    ws.append(
        [
            "TicketID",
            "Erstellt",
            "Gerät/Modell",
            "Problem Kurz",
            "Problem Lang",
            "Lösung",
            "Fehlercodes",
            "Ersatzteile",
            "Reparaturzeit (min)",
        ]
    )

    # Sample data
    base_date = datetime.now()
    tickets = [
        [
            f"OF-{1000 + i}",
            (base_date - timedelta(days=i)).isoformat(),
            f"KM C{308 + i*10}",
            f"Paper jam issue {i}",
            f"Customer reported paper jam in input tray {i}",
            f"Cleared jam and reset counter {i}",
            "13.B9;50.FF",
            "RM1-0542-000;FK3-2105-000",
            15,
        ]
        for i in range(100)
    ]

    for ticket in tickets:
        ws.append(ticket)

    # Write to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@pytest.fixture
def mock_km_excel_pp():
    """Create mock KM Production (PP) Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tickets"

    ws.append(
        [
            "TicketID",
            "Erstellt",
            "Gerät/Modell",
            "Problem Kurz",
            "Problem Lang",
            "Lösung",
            "Fehlercodes",
            "Ersatzteile",
            "Reparaturzeit (min)",
        ]
    )

    base_date = datetime.now()
    for i in range(50):
        ws.append(
            [
                f"PP-{2000 + i}",
                (base_date - timedelta(days=i * 2)).isoformat(),
                f"KM C{557 + (i % 5)*10}",
                f"Toner low {i}",
                f"Toner cartridge running low {i}",
                f"Replaced toner {i}",
                "11.A3",
                "TN-321",
                5,
            ]
        )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@pytest.fixture
def mock_km_excel_sol():
    """Create mock KM Solutions (SOL) Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tickets"

    ws.append(
        [
            "TicketID",
            "Erstellt",
            "Gerät/Modell",
            "Problem Kurz",
            "Problem Lang",
            "Lösung",
            "Fehlercodes",
            "Ersatzteile",
            "Reparaturzeit (min)",
        ]
    )

    base_date = datetime.now()
    for i in range(30):
        ws.append(
            [
                f"SOL-{3000 + i}",
                (base_date - timedelta(days=i * 3)).isoformat(),
                f"KM C{308}",
                f"Network issue {i}",
                f"Device not connecting to network {i}",
                f"Reconfigured network settings {i}",
                "",
                "",
                30,
            ]
        )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@pytest.fixture
def mock_km_part_lifetimes():
    """Create mock KM part lifetimes (consumables) Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PartLifetimes"

    ws.append(
        [
            "Manufacturer",
            "ModelFamily",
            "PartCategory",
            "PartNumber",
            "NominalLifetimePages",
            "ColorChannel",
        ]
    )

    # Realistic part lifetime data for various model families
    model_families = ["ZEUS", "EAGLE", "FALCON", "C658", "C558", "C458"]
    parts_data = [
        ("toner", "TN-514", 40000, "K"),
        ("drum", "DR-512", 200000, "K"),
        ("fuser", "FK3-2105", 600000, None),
        ("transfer_belt", "TB-01", 400000, None),
        ("pickup_roller", "PR-001", 100000, None),
    ]

    for model_idx, model in enumerate(model_families):
        for part_cat, part_num, lifetime, color in parts_data:
            if color:  # Toners are per-color
                for c in ["K", "C", "M", "Y"]:
                    ws.append(
                        [
                            "Konica Minolta, Inc.",
                            model,
                            part_cat,
                            f"{part_num}-{c}",
                            lifetime * (0.9 + 0.1 * model_idx),
                            c,
                        ]
                    )
            else:
                ws.append(["Konica Minolta, Inc.", model, part_cat, part_num, lifetime * (0.9 + 0.1 * model_idx), None])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
