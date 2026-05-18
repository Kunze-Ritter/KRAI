"""Unit tests for PartLifetimesImporter."""

from unittest.mock import AsyncMock

import pytest

from backend.processors.part_lifetimes_importer import PartLifetimesImporter


@pytest.fixture
def mock_db_adapter():
    """Create mock database adapter."""
    adapter = AsyncMock()
    adapter.execute = AsyncMock(return_value=None)
    adapter.fetch_one = AsyncMock()
    return adapter


@pytest.fixture
def importer(mock_db_adapter):
    """Create importer instance."""
    return PartLifetimesImporter(mock_db_adapter)


@pytest.mark.asyncio
async def test_ingest_part_lifetimes_success(importer, mock_km_part_lifetimes, mock_db_adapter):
    """Test successful ingestion of part lifetimes."""

    # Mock manufacturer lookup to return valid IDs
    async def mock_fetch_one(sql, params):
        if "manufacturers" in sql:
            # Return a fake manufacturer UUID
            return ("550e8400-e29b-41d4-a716-446655440000",)
        return None

    mock_db_adapter.fetch_one = mock_fetch_one
    importer.db_adapter = mock_db_adapter

    count = await importer.ingest_file(mock_km_part_lifetimes)

    assert count > 0


@pytest.mark.asyncio
async def test_ingest_missing_excel_file(importer):
    """Test that missing file raises error."""
    with pytest.raises(ValueError) as exc_info:
        await importer.ingest_file(None)

    assert "Missing excel_file" in str(exc_info.value)


@pytest.mark.asyncio
async def test_extract_part_lifetimes(importer, mock_km_part_lifetimes):
    """Test extraction of part lifetime entries."""
    import openpyxl

    workbook = openpyxl.load_workbook(mock_km_part_lifetimes, data_only=True)
    worksheet = workbook.active

    entries = await importer._extract_part_lifetimes(worksheet)

    # Should extract multiple entries
    assert len(entries) > 0

    # Check structure of first entry
    first_entry = entries[0]
    assert "manufacturer" in first_entry
    assert "part_category" in first_entry
    assert "nominal_lifetime_pages" in first_entry
    assert isinstance(first_entry["nominal_lifetime_pages"], int)
    assert first_entry["nominal_lifetime_pages"] > 0


@pytest.mark.asyncio
async def test_part_category_normalization(importer):
    """Test that part categories are properly normalized."""
    from io import BytesIO

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Manufacturer", "ModelFamily", "PartCategory", "PartNumber", "NominalLifetimePages", "ColorChannel"])

    test_categories = ["TONER", "Drum", "FUSER", "Transfer_Belt"]
    for i, cat in enumerate(test_categories):
        ws.append(["Konica Minolta", "C308", cat, f"PN{i}", "50000", "K"])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    workbook = openpyxl.load_workbook(output, data_only=True)
    worksheet = workbook.active
    entries = await importer._extract_part_lifetimes(worksheet)

    # Check that categories are lowercase
    for entry, original_cat in zip(entries, test_categories):
        assert entry["part_category"] == original_cat.lower()


@pytest.mark.asyncio
async def test_lifetime_pages_parsing(importer):
    """Test various lifetime page number formats."""
    from io import BytesIO

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Manufacturer", "ModelFamily", "PartCategory", "PartNumber", "NominalLifetimePages", "ColorChannel"])
    ws.append(["Konica Minolta", "C308", "toner", "PN001", "50000", "K"])
    ws.append(["Konica Minolta", "C308", "drum", "PN002", "200000.0", "K"])  # Float
    ws.append(["Konica Minolta", "C308", "fuser", "PN003", "600000.5", None])  # Float with decimal

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    workbook = openpyxl.load_workbook(output, data_only=True)
    worksheet = workbook.active
    entries = await importer._extract_part_lifetimes(worksheet)

    assert entries[0]["nominal_lifetime_pages"] == 50000
    assert entries[1]["nominal_lifetime_pages"] == 200000
    assert entries[2]["nominal_lifetime_pages"] == 600000  # Float truncated to int


@pytest.mark.asyncio
async def test_optional_model_family(importer):
    """Test that ModelFamily is optional."""
    from io import BytesIO

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Manufacturer", "ModelFamily", "PartCategory", "PartNumber", "NominalLifetimePages", "ColorChannel"])
    ws.append(["Konica Minolta", "C308", "toner", "PN001", "50000", "K"])
    ws.append(["Konica Minolta", None, "drum", "PN002", "200000", None])  # No model family

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    workbook = openpyxl.load_workbook(output, data_only=True)
    worksheet = workbook.active
    entries = await importer._extract_part_lifetimes(worksheet)

    assert len(entries) >= 2
    assert entries[0]["model_family"] == "C308"
    assert entries[1]["model_family"] is None


def test_known_part_categories(importer):
    """Test that processor knows all standard part categories."""
    assert "toner" in importer.PART_CATEGORIES
    assert "drum" in importer.PART_CATEGORIES
    assert "fuser" in importer.PART_CATEGORIES
    assert "transfer_belt" in importer.PART_CATEGORIES
